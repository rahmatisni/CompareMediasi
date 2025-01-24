import json
import mysql.connector
import signal
from datetime import date, datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import PatternFill

card_number = [
    '6032982888011432',
    '6032981030212971',
    '6032982601347550',
    '6032982838195285',
    '6032982868891191',
    '0145008201654190',
    '6032984041529904',
    '0145200036176123',
    '6032986098872103',
    '0145008401176242',
    '0145200305670848',
    '6032982815512775',
    '0145201600543706',
    '0145200024277529',
    '6032984065346391',
    '6013500124405284',
    '6032982540917745',
    '7546060001888098',
    '0145008400868906',
    '6032982846512513',
    '6032982851815538',
    '6032984099676979',
    '0145015300110149',
    '0145008201654166',
    '6032984008709101',
    '6013500438019110',
    '6032982862303110',
    '6032986082866459'
]

with open('mediasi.json', 'r') as f:
    credentials = json.load(f)

today_date = datetime.now().strftime("%Y%m%d")  # Format: YYYYMMDD

# Timeout handler function


def timeout_handler(signum, frame):
    raise TimeoutError("Operation timed out")

# Function to recursively convert values for JSON serialization


def convert_for_json(data):
    if isinstance(data, list):
        return [convert_for_json(item) for item in data]
    elif isinstance(data, dict):
        return {key: convert_for_json(value) for key, value in data.items()}
    elif isinstance(data, datetime):
        # Format lengkap dengan jam, menit, dan detik
        return data.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(data, date):
        return data.strftime("%Y-%m-%d")
    elif isinstance(data, Decimal):
        return float(data)
    elif isinstance(data, bytes):
        return data.decode('utf-8')
    else:
        return data


# Define Excel styles
green_fill = PatternFill(start_color="00FF00",
                         end_color="00FF00", fill_type="solid")
light_green_fill = PatternFill(
    start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
blue_fill = PatternFill(start_color="0000FF",
                        end_color="0000FF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00",
                          end_color="FFFF00", fill_type="solid")

purple_fill = PatternFill(start_color="800080", 
                          end_color="800080", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", 
                          end_color="FFA500", fill_type="solid")
pink_fill = PatternFill(start_color="FFC0CB", 
                        end_color="FFC0CB", fill_type="solid")

# Function to write JSON data to Excel


def write_to_excel(workbook, sheet_name, data):
    sheet = workbook.create_sheet(sheet_name)
    if not data:
        sheet.append(["No Data"])
        return

    # Write header
    headers = data[0].keys()
    sheet.append(list(headers))

    # Write data rows
    for row in data:
        values = list(row.values())
        sheet.append(values)

# Highlight rows based on three sheets comparison
# def highlight_rows_three_sheets(workbook, sheet1_name, sheet2_name, sheet3_name, date_column_name="tgl_transaksi"):
#     sheet1 = workbook[sheet1_name]
#     sheet2 = workbook[sheet2_name]
#     sheet3 = workbook[sheet3_name]

#     sheet1_dates = set()
#     sheet2_dates = set()
#     sheet3_dates = set()

#     # Extract headers
#     sheet1_header = [cell.value for cell in sheet1[1]]
#     sheet2_header = [cell.value for cell in sheet2[1]]
#     sheet3_header = [cell.value for cell in sheet3[1]]

#     # Find column indexes
#     date_col_idx_sheet1 = sheet1_header.index(date_column_name) + 1 if date_column_name in sheet1_header else None
#     date_col_idx_sheet2 = sheet2_header.index(date_column_name) + 1 if date_column_name in sheet2_header else None
#     date_col_idx_sheet3 = sheet3_header.index(date_column_name) + 1 if date_column_name in sheet3_header else None

#     # Collect dates
#     if date_col_idx_sheet1:
#         sheet1_dates = {str(row[date_col_idx_sheet1 - 1].value) for row in sheet1.iter_rows(min_row=2) if row[date_col_idx_sheet1 - 1].value}
#     if date_col_idx_sheet2:
#         sheet2_dates = {str(row[date_col_idx_sheet2 - 1].value) for row in sheet2.iter_rows(min_row=2) if row[date_col_idx_sheet2 - 1].value}
#     if date_col_idx_sheet3:
#         sheet3_dates = {str(row[date_col_idx_sheet3 - 1].value) for row in sheet3.iter_rows(min_row=2) if row[date_col_idx_sheet3 - 1].value}

#     # Determine overlaps
#     overlap_all = sheet1_dates & sheet2_dates & sheet3_dates
#     overlap_1_2 = (sheet1_dates & sheet2_dates) - overlap_all
#     overlap_2_3 = (sheet2_dates & sheet3_dates) - overlap_all
#     overlap_1_3 = (sheet1_dates & sheet3_dates) - overlap_all

#     # Highlight rows
#     def highlight(sheet, dates, col_idx, fill):
#         for row in sheet.iter_rows(min_row=2):
#             if str(row[col_idx - 1].value) in dates:
#                 for cell in row:
#                     cell.fill = fill

#     if date_col_idx_sheet1:
#         highlight(sheet1, overlap_all, date_col_idx_sheet1, green_fill)
#         highlight(sheet1, overlap_1_2, date_col_idx_sheet1, blue_fill)
#         highlight(sheet1, overlap_1_3, date_col_idx_sheet1, yellow_fill)

#     if date_col_idx_sheet2:
#         highlight(sheet2, overlap_all, date_col_idx_sheet2, green_fill)
#         highlight(sheet2, overlap_1_2, date_col_idx_sheet2, blue_fill)
#         highlight(sheet2, overlap_2_3, date_col_idx_sheet2, light_green_fill)
#     if date_col_idx_sheet3:
#         highlight(sheet3, overlap_all, date_col_idx_sheet3, green_fill)
#         highlight(sheet3, overlap_1_3, date_col_idx_sheet3, yellow_fill)
#         highlight(sheet3, overlap_2_3, date_col_idx_sheet3, light_green_fill)


def highlight_rows_three_sheets(workbook, sheet1_name, sheet2_name, sheet3_name,sheet4_name, date_column_name="tgl_transaksi", card_column_name="no_kartu"):
    sheet1 = workbook[sheet1_name]
    sheet2 = workbook[sheet2_name]
    sheet3 = workbook[sheet3_name]
    sheet4 = workbook[sheet4_name]


    sheet1_entries = set()
    sheet2_entries = set()
    sheet3_entries = set()
    sheet4_entries = set()

    # Extract headers
    sheet1_header = [cell.value for cell in sheet1[1]]
    sheet2_header = [cell.value for cell in sheet2[1]]
    sheet3_header = [cell.value for cell in sheet3[1]]
    sheet4_header = [cell.value for cell in sheet4[1]]


    # Find column indexes
    date_col_idx_sheet1 = sheet1_header.index(
        date_column_name) + 1 if date_column_name in sheet1_header else None
    card_col_idx_sheet1 = sheet1_header.index(
        card_column_name) + 1 if card_column_name in sheet1_header else None

    date_col_idx_sheet2 = sheet2_header.index(
        date_column_name) + 1 if date_column_name in sheet2_header else None
    card_col_idx_sheet2 = sheet2_header.index(
        card_column_name) + 1 if card_column_name in sheet2_header else None

    date_col_idx_sheet3 = sheet3_header.index(
        date_column_name) + 1 if date_column_name in sheet3_header else None
    card_col_idx_sheet3 = sheet3_header.index(
        card_column_name) + 1 if card_column_name in sheet3_header else None
    
    date_col_idx_sheet4 = sheet4_header.index(
        date_column_name) + 1 if date_column_name in sheet4_header else None
    card_col_idx_sheet4 = sheet4_header.index(
        card_column_name) + 1 if card_column_name in sheet4_header else None

    # Collect entries (date and card number)
    if date_col_idx_sheet1 and card_col_idx_sheet1:
        sheet1_entries = {(str(row[date_col_idx_sheet1 - 1].value), str(row[card_col_idx_sheet1 - 1].value))
                          for row in sheet1.iter_rows(min_row=2) if row[date_col_idx_sheet1 - 1].value and row[card_col_idx_sheet1 - 1].value}

    if date_col_idx_sheet2 and card_col_idx_sheet2:
        sheet2_entries = {(str(row[date_col_idx_sheet2 - 1].value), str(row[card_col_idx_sheet2 - 1].value))
                          for row in sheet2.iter_rows(min_row=2) if row[date_col_idx_sheet2 - 1].value and row[card_col_idx_sheet2 - 1].value}

    if date_col_idx_sheet3 and card_col_idx_sheet3:
        sheet3_entries = {(str(row[date_col_idx_sheet3 - 1].value), str(row[card_col_idx_sheet3 - 1].value))
                          for row in sheet3.iter_rows(min_row=2) if row[date_col_idx_sheet3 - 1].value and row[card_col_idx_sheet3 - 1].value}
    if date_col_idx_sheet4 and card_col_idx_sheet4:
        sheet4_entries = {(str(row[date_col_idx_sheet4 - 1].value), str(row[card_col_idx_sheet4 - 1].value))
                          for row in sheet4.iter_rows(min_row=2) if row[date_col_idx_sheet4 - 1].value and row[card_col_idx_sheet4 - 1].value}

# Determine overlaps
    overlap_all = sheet1_entries & sheet2_entries & sheet3_entries & sheet4_entries
# Overlap for combinations of 2 sheets
    overlap_1_2 = (sheet1_entries & sheet2_entries) - overlap_all
    overlap_1_3 = (sheet1_entries & sheet3_entries) - overlap_all
    overlap_1_4 = (sheet1_entries & sheet4_entries) - overlap_all
    overlap_2_3 = (sheet2_entries & sheet3_entries) - overlap_all
    overlap_2_4 = (sheet2_entries & sheet4_entries) - overlap_all
    overlap_3_4 = (sheet3_entries & sheet4_entries) - overlap_all

    # Overlap for combinations of 3 sheets
    overlap_1_2_3 = (sheet1_entries & sheet2_entries & sheet3_entries) - overlap_all
    overlap_1_2_4 = (sheet1_entries & sheet2_entries & sheet4_entries) - overlap_all
    overlap_1_3_4 = (sheet1_entries & sheet3_entries & sheet4_entries) - overlap_all
    overlap_2_3_4 = (sheet2_entries & sheet3_entries & sheet4_entries) - overlap_all


    # Unique entries
    # Determine unique entries for each sheet
    unique_sheet1 = sheet1_entries - (sheet2_entries | sheet3_entries | sheet4_entries)
    unique_sheet2 = sheet2_entries - (sheet1_entries | sheet3_entries | sheet4_entries)
    unique_sheet3 = sheet3_entries - (sheet1_entries | sheet2_entries | sheet4_entries)
    unique_sheet4 = sheet4_entries - (sheet1_entries | sheet2_entries | sheet3_entries)


    # Highlight rows
    def highlight(sheet, entries, date_col_idx, card_col_idx, fill):
        for row in sheet.iter_rows(min_row=2):
            entry = (str(row[date_col_idx - 1].value),
                    str(row[card_col_idx - 1].value))
            if entry in entries:
                for cell in row:
                    cell.fill = fill

    # Highlight entries for sheet1
    # Highlight entries for sheet1


    # Colors for highlighting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Unique entries
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Overlap all sheets
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Overlap 1 & 2
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Overlap 1 & 3
    purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Overlap 1 & 4
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Overlap 2 & 3
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Overlap 2 & 4
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Overlap 3 & 4

    # Fill colors for 3-sheet overlaps
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Overlap 1, 2 & 3
    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Overlap 1, 2 & 4
    light_purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Overlap 1, 3 & 4
    light_pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Overlap 2, 3 & 4


    if date_col_idx_sheet1 and card_col_idx_sheet1:
        highlight(sheet1, overlap_all, date_col_idx_sheet1, card_col_idx_sheet1, green_fill)
        highlight(sheet1, overlap_1_2, date_col_idx_sheet1, card_col_idx_sheet1, blue_fill)
        highlight(sheet1, overlap_1_3, date_col_idx_sheet1, card_col_idx_sheet1, yellow_fill)
        highlight(sheet1, overlap_1_4, date_col_idx_sheet1, card_col_idx_sheet1, purple_fill)
        highlight(sheet1, overlap_1_2_3, date_col_idx_sheet1, card_col_idx_sheet1, light_blue_fill)
        highlight(sheet1, overlap_1_2_4, date_col_idx_sheet1, card_col_idx_sheet1, light_yellow_fill)
        highlight(sheet1, overlap_1_3_4, date_col_idx_sheet1, card_col_idx_sheet1, light_purple_fill)
        highlight(sheet1, unique_sheet1, date_col_idx_sheet1, card_col_idx_sheet1, red_fill)

    # Highlight entries for sheet2
    if date_col_idx_sheet2 and card_col_idx_sheet2:
        highlight(sheet2, overlap_all, date_col_idx_sheet2, card_col_idx_sheet2, green_fill)
        highlight(sheet2, overlap_1_2, date_col_idx_sheet2, card_col_idx_sheet2, blue_fill)
        highlight(sheet2, overlap_2_3, date_col_idx_sheet2, card_col_idx_sheet2, light_green_fill)
        highlight(sheet2, overlap_2_4, date_col_idx_sheet2, card_col_idx_sheet2, orange_fill)
        highlight(sheet2, overlap_1_2_3, date_col_idx_sheet2, card_col_idx_sheet2, light_blue_fill)
        highlight(sheet2, overlap_1_2_4, date_col_idx_sheet2, card_col_idx_sheet2, light_yellow_fill)
        highlight(sheet2, overlap_2_3_4, date_col_idx_sheet2, card_col_idx_sheet2, light_pink_fill)
        highlight(sheet2, unique_sheet2, date_col_idx_sheet2, card_col_idx_sheet2, red_fill)

    # Highlight entries for sheet3
    if date_col_idx_sheet3 and card_col_idx_sheet3:
        highlight(sheet3, overlap_all, date_col_idx_sheet3, card_col_idx_sheet3, green_fill)
        highlight(sheet3, overlap_1_3, date_col_idx_sheet3, card_col_idx_sheet3, yellow_fill)
        highlight(sheet3, overlap_2_3, date_col_idx_sheet3, card_col_idx_sheet3, light_green_fill)
        highlight(sheet3, overlap_3_4, date_col_idx_sheet3, card_col_idx_sheet3, pink_fill)
        highlight(sheet3, overlap_1_2_3, date_col_idx_sheet3, card_col_idx_sheet3, light_blue_fill)
        highlight(sheet3, overlap_1_3_4, date_col_idx_sheet3, card_col_idx_sheet3, light_purple_fill)
        highlight(sheet3, overlap_2_3_4, date_col_idx_sheet3, card_col_idx_sheet3, light_pink_fill)
        highlight(sheet3, unique_sheet3, date_col_idx_sheet3, card_col_idx_sheet3, red_fill)

    # Highlight entries for sheet4
    if date_col_idx_sheet4 and card_col_idx_sheet4:
        highlight(sheet4, overlap_all, date_col_idx_sheet4, card_col_idx_sheet4, green_fill)
        highlight(sheet4, overlap_1_4, date_col_idx_sheet4, card_col_idx_sheet4, purple_fill)
        highlight(sheet4, overlap_2_4, date_col_idx_sheet4, card_col_idx_sheet4, orange_fill)
        highlight(sheet4, overlap_3_4, date_col_idx_sheet4, card_col_idx_sheet4, pink_fill)
        highlight(sheet4, overlap_1_2_4, date_col_idx_sheet4, card_col_idx_sheet4, light_yellow_fill)
        highlight(sheet4, overlap_1_3_4, date_col_idx_sheet4, card_col_idx_sheet4, light_purple_fill)
        highlight(sheet4, overlap_2_3_4, date_col_idx_sheet4, card_col_idx_sheet4, light_pink_fill)
        highlight(sheet4, unique_sheet4, date_col_idx_sheet4, card_col_idx_sheet4, red_fill)


    summary_sheet = workbook.create_sheet(title="Summary")
    summary_sheet.append(["Sheet", "Total Records", "Data Unique", "Info", "Persentase"])
    summary_sheet.append([sheet1_name, sheet1.max_row - 1, len(unique_sheet1), f"-"])
    summary_sheet.append([sheet2_name, sheet2.max_row - 1, len(unique_sheet2),"Akurasi Mediasi DB HIST", f"{(sheet2.max_row - 1)/(len(unique_sheet1)+len(unique_sheet3)+len(unique_sheet4)+(sheet2.max_row - 1)) * 100:.2f}%"])
    summary_sheet.append([sheet3_name, sheet3.max_row - 1, len(unique_sheet3), "Akurasi DB SMT", f"{(sheet3.max_row - 1)/(len(unique_sheet1)+len(unique_sheet2)+len(unique_sheet4)+(sheet3.max_row - 1)) * 100:.2f}%"])
    summary_sheet.append([sheet4_name, sheet4.max_row - 1, len(unique_sheet4),"Akurasi Mediasi DB RUAS", f"{(sheet4.max_row - 1)/(len(unique_sheet1)+len(unique_sheet2)+len(unique_sheet3)+(sheet4.max_row - 1)) * 100:.2f}%"])

# Main execution
# card_number = input("Masukkan nomor kartu:
# Daftar nomor kartu

# Menambahkan data baru dengan menghapus '0' di awal
new_card_numbers = [card[1:] for card in card_number if card.startswith('0')]

# Menggabungkan dengan daftar yang sudah ada
card_number.extend(new_card_numbers)

# Menampilkan hasil
print(card_number)
card_number_str = ', '.join(f"'{cn}'" for cn in card_number)

# Primary connection
conn = mysql.connector.connect(
    host="172.16.4.8",
    user="jmto",
    password="@jmt02024!#",
    connection_timeout=10
)

query1_results = []
query2_results = []
query3_results = []
query4_results = []


try:
    cursor = conn.cursor()

    # Get databases
    cursor.execute("SHOW DATABASES;")
    all_databases = cursor.fetchall()
    databases = [db[0] for db in all_databases if 'lattol' in db[0]]

    if not databases:
        print("No databases found containing 'lattol'")
    else:
        for db_name in databases:
            print(f"Checking database Copy: {db_name}")
            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(600)
            try:
                # #
                #                 query = f"""
                #                     SELECT * FROM {db_name}.jid_transaksi_deteksi
                #                     WHERE (etoll_id = '{card_number}' OR etoll_id = {int(card_number)})
                #                     AND tgl_transaksi >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
                #                     AND tgl_transaksi <= CURDATE();
                #                 """

                # query = f"""
                #     SELECT *, etoll_id AS nomor_kartu
                #     FROM {db_name}.jid_transaksi_deteksi
                #     WHERE etoll_id IN ({card_number_str})
                #     AND tgl_transaksi >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
                #     AND tgl_transaksi <= CURDATE();
                # """
                query = f"""
                    SELECT
                    id,
                        etoll_id as 'no_kartu',
                        ruas_id,
                        asal_gerbang_id,
                        gerbang_id,
                        gardu_id,
                        tgl_lap,
                        shift,
                        perioda,
                        no_resi,
                        gol_sah,
                        metoda_bayar_sah,
                        jenis_notran,
                        validasi_notran,
                        tgl_transaksi,
                        kspt_id,
                        pultol_id,
                        tgl_entrance,
                        id_obu,
                        etoll_hash,
                        tarif,
                        sisa_saldo
                        FROM {db_name}.jid_transaksi_deteksi
                        WHERE etoll_id IN ({card_number_str})
                        and tarif != 0 and etoll_hash != 0
                        AND tgl_transaksi >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
                        AND tgl_transaksi <= CURDATE();
                    """

                cursor.execute(query)
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()
                if results:
                    query1_results.extend(
                        [convert_for_json(dict(zip(columns, row))) for row in results])
            except TimeoutError:
                print(f"Timeout: Query for database {db_name} took too long.")
            except mysql.connector.Error as query_err:
                print(f"Query error in database {db_name}: {query_err}")
            finally:
                signal.alarm(0)

    # Query 2
    cursor.execute(f"""
    SELECT * FROM travoy_db_history.tx_card_toll_history
    WHERE no_kartu IN ({card_number_str})
 AND tgl_transaksi >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
  AND tgl_transaksi <= CURDATE();
    """)

    print(f"Checking database: {db_name}")

    columns = [col[0] for col in cursor.description]
    query2_results.extend([convert_for_json(dict(zip(columns, row)))
                          for row in cursor.fetchall()])

finally:
    cursor.close()
    conn.close()

# Separate connection for Query 3
try:
    conn_additional = mysql.connector.connect(
        host="20.10.20.99",
        user="travoyhis",
        password="DbHistoryTravoy2021!",
        connection_timeout=10
    )
    cursor_additional = conn_additional.cursor()

    cursor_additional.execute(f"""
        SELECT * FROM travoy_db_history.tx_card_toll_history
                        WHERE no_kartu IN ({card_number_str})
        AND tgl_transaksi >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
        AND tgl_transaksi <= CURDATE()
        GROUP BY no_kartu, tgl_transaksi;
    """)

    print(f"Checking database: {'travoy_db_history'}")

    columns = [col[0] for col in cursor_additional.description]
    query3_results.extend([convert_for_json(dict(zip(columns, row)))
                          for row in cursor_additional.fetchall()])

finally:
    cursor_additional.close()
    conn_additional.close()


for credential in credentials:
    print(
        f"Connecting to database: {credential['host']} as {credential['user_name']}")
    try:
        conn = mysql.connector.connect(
            host=credential['host'],
            user=credential['user_name'],
            password=credential['password'],
            database=credential.get('default_database', None)
        )

        cursor = conn.cursor()

        # Get databases
        cursor.execute("SHOW DATABASES;")
        all_databases = cursor.fetchall()
        databases = [db[0] for db in all_databases if 'lattol' in db[0]]

        if not databases:
            print("No databases found containing 'lattol'")
        else:
            for db_name in databases:
                print(f"Checking database Ruas: {db_name}")
                signal.signal(signal.SIGALRM, timeout_handler)
                signal.alarm(600)
                try:
                    #                     query = f"""
                    #                     SELECT * FROM {db_name}.jid_transaksi_deteksi
                    #                     WHERE etoll_id = '{card_number}' AND tgl_transaksi >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
                    #   AND tgl_transaksi <= CURDATE();
                    #                     """

                    query = f"""
                    SELECT
                    id,
                        etoll_id as 'no_kartu',
                        ruas_id,
                        asal_gerbang_id,
                        gerbang_id,
                        gardu_id,
                        tgl_lap,
                        shift,
                        perioda,
                        no_resi,
                        gol_sah,
                        metoda_bayar_sah,
                        jenis_notran,
                        validasi_notran,
                        tgl_transaksi,
                        kspt_id,
                        pultol_id,
                        tgl_entrance,
                        id_obu,
                        etoll_hash,
                        tarif,
                        sisa_saldo
                        FROM {db_name}.jid_transaksi_deteksi
                        WHERE etoll_id IN ({card_number_str})
                        and tarif != 0 and etoll_hash != 0
                        AND tgl_transaksi >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
                        AND tgl_transaksi <= CURDATE();
                    """
                    print('querynya', query)
                    cursor.execute(query)
                    columns = [col[0] for col in cursor.description]
                    results = cursor.fetchall()
                    if results:
                        query4_results.extend(
                            [convert_for_json(dict(zip(columns, row))) for row in results])
                except TimeoutError:
                    print(
                        f"Timeout: Query for database {db_name} took too long.")
    except mysql.connector.Error as query_err:
        print(f"Query error in database {db_name}: {query_err}")
    finally:
        signal.alarm(0)


# Create and save workbook
workbook = Workbook()
workbook.remove(workbook.active)

write_to_excel(workbook, "Mediasi Copy", query1_results)
write_to_excel(workbook, "Mediasi DB Hist", query2_results)
write_to_excel(workbook, "SMT DB Hist", query3_results)
write_to_excel(workbook, "Mediasi DB RUAS", query4_results)

highlight_rows_three_sheets(
    workbook, "Mediasi Copy", "Mediasi DB Hist", "SMT DB Hist","Mediasi DB RUAS")

# output_file = f"{card_number}_{today_date}_compare.xlsx"
output_file = f"{today_date}_compare.xlsx"

workbook.save(output_file)
print(f"Nomor Kartu {card_number}")

print(f"Excel file saved as {output_file}")
